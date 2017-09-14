# -*- coding: utf-8 -*-
from bs4 import BeautifulSoup
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import Bootstrap

str_1 = Bootstrap.str_1	#	Dedicated to Z.R.H. Ich liebe dich.
cik = Bootstrap.cik
year = Bootstrap.year
types = Bootstrap.types
if types in ['10-Q', '10-q']:
	quart = Bootstrap.quart

def typeset(ncols, nrows, sheet):	#	This function is used to eliminate the blank columns.
	outlier_col = []	#	To get to know which columns are blank.
	for j in range(1, ncols + 1):
		flag_1 = 0
		for k in range(1, nrows + 1):
			var = sheet.cell(row = k, column = j).value
			if var != '':
				flag_1 = 1
		if flag_1 == 0:
			outlier_col.append(j)

	well_col = []	# To get to know which columns have been filled with.
	for i in range(1, ncols + 1):
		if not i in outlier_col:
			well_col.append(i)

	for j in range(1, nrows + 1):	#	To move those columns which are not blank.
		for k in range(1, len(well_col) + 1):
			sheet.cell(row = j, column = k).value = sheet.cell(row = j, column = well_col[k - 1]).value
			if sheet.cell(row = j, column = well_col[k - 1]).font == font_1:
				sheet.cell(row = j, column = k).font = font_1
			else:
				sheet.cell(row = j, column = k).font = font_2
			if well_col[k - 1] > k:
				sheet.cell(row = j, column = well_col[k - 1]).value = ''
	return len(well_col)

def arrange(ncols, nrows, title, sheet, first_list):
	if ncols > 2:
		title_cols = []
		for k in range(ncols + 1, 1, -1):
			if sheet.cell(row = title, column = k).value != '':
				title_cols.append(k)
		if title_cols == []:
			title_cols = [2]
		title_col = []
		for k in range(2, ncols + 1):
			if sheet.cell(row = title, column = k).value != '':
				title_col.append(k)
		if title_col == []:
			title_col = [2]
		sort_list = list(range(2, ncols + 1))
		sort_list.reverse()
		for cols in sort_list:
			if sheet.cell(row = title, column = cols).value != '':
				break
		col_flag = 0
		for col in range(2, ncols + 1):
			if sheet.cell(row = title, column = col).value == '':
				col_flag = 1
				break
		super = 0
		for unit_1 in first_list:
			jump_flag = 0
			for unit_0 in title_col:
				if sheet.cell(row = unit_1, column = unit_0).value == '':
					if (sheet.cell(row = unit_1, column = unit_0 + 1).value != '' and not unit_0 + 1 in title_col):	#	Dedicated to Z.R.H. Ich liebe dich.
						super = unit_0 + 1
						jump_flag = 1
						break
			if jump_flag == 1:
				break
		if (col_flag == 1 and super != 0):
			move_rows = []
			for rows in range(title + 1, nrows + 1):
				if sheet.cell(row = rows, column = super).value != '':
					move_rows.append(rows)
			book = []
			for k in range(super, ncols + 1):
				if k <= title_cols[0]:
					for j in move_rows:
						if sheet.cell(row = j, column = k).value != '':
							book.append(k)
							break
				else:
					if k > title_cols[0]:
						for j in range(title + 1, nrows + 1):
							if sheet.cell(row = j, column = k).value != '':
								book.append(k)
								break
			if move_rows != []:
				for d in book:
					sym_flag = 0
					for k in title_cols:
						if k < d:
							if sheet.cell(row = move_rows[0], column = k).value == '':
								for j in move_rows:
									sheet.cell(row = j, column = k).value = sheet.cell(row = j, column = d).value
									sheet.cell(row = j, column = d).value = ''
									sym_flag = 1
						if sym_flag == 1:
							break

print('**********************Phase II : Capturing financial statements existing in the report**********************')
soup = BeautifulSoup(str_1,'lxml')
table_list = soup.find_all(name = 'table')
font_1 = Font(name = 'Times New Roman', size = 11, bold = False, italic = False, vertAlign = None, underline = 'none', strike=False, color='FF000000')
font_2 = Font(name = 'Times New Roman', size = 11, bold = True, italic = False, vertAlign = None, underline = 'none', strike=False, color='FF000000')
align_1 = Alignment(horizontal = 'left', vertical = 'bottom', text_rotation = 0, wrap_text = False, shrink_to_fit = False, indent = 1)
fill_1 = PatternFill(fill_type = 'solid', start_color = 'CCEEFF', end_color = 'CCEEFF')
w = Workbook()

print('**********************Phase III : Generating XLSX**********************')
comp = re.compile('<table[\s\S]+?>[\s\S]+?</table>', re.I)
comp_0 = re.compile('cceeff', re.I)
mirror = re.findall(comp, str_1)	#	Prepare for removing those pages with discrepancy.
out_list = []
for term in range(len(mirror)):
	obj = mirror[term]
	if re.search(comp_0, obj) is None:
		out_list.append(term)
tab_list = []	#	Remove those pages with discrepancy.
for term in range(len(table_list)):
	con = table_list[term].contents
	if not term in out_list:
		tab_list.append(table_list[term])
for tab in tab_list:
	sheet = w.create_sheet()
	tr_list = tab.find_all(name = 'tr')
	row_0 = 0
	ratio_set = []
	margin_flag = 0
	indent_flags = 0
	for item in tr_list:
		row_0 += 1
		cont = item.find_all(name = 'td')
		col_0 = 0
		col_sum = 0
		for term in cont:
			col_0 += 1
			child_0 = term.descendants
			bold_flag = 0
			indent_flag = 0
			groups = []
			flag_0 = 0
			for doc in child_0:
				if doc != '\n':
					if not doc.string is None:
						string_0 = doc.string
						string_1 = string_0.strip()
						word = []
						flag_0 = 1
						if len(string_1.splitlines()) != 1:	#	Settle the problem of NEWLINES.
							for voc in string_1.splitlines():
								voc = voc.strip()
								word.append(voc)
							string_1 = ' '.join(word)
						groups.append(string_1)
						if doc.name == 'b':
							bold_flag = 1
				if (doc.name == 'font' and 'style' in list(doc.attrs.keys())):
					if 'font-weight:bold' in doc['style']:
						bold_flag = 1
				if (col_0 == 1 and not term.name is None and indent_flags == 0): 
					if	'style' in list(term.attrs.keys()):
						if 'padding-left:' in term['style']:
							padding_norm = re.search('padding-left:[\s\d\.-]+?', term['style']).group()
							indent_flags = 1
				if (col_0 == 1 and not term.name is None and indent_flags == 1): 
					if	'style' in list(term.attrs.keys()):
						if 'padding-left:' in term['style']:
							padding = re.search('padding-left:[\s\d\.-]+?', term['style']).group()
							if padding != padding_norm:
								indent_flag = 1
				if (col_0 == 1 and not doc.name is None and margin_flag == 0): 
					if	'style' in list(doc.attrs.keys()):
						if 'margin-left' in doc['style']:
							margin_norm = re.search('margin-left:[\s\d\.-]+?', doc['style']).group()
							margin_flag = 1
				if (col_0 == 1 and not doc.name is None and margin_flag == 1): 
					if	'style' in list(doc.attrs.keys()):
						if 'margin-left' in doc['style']:
							margin = re.search('margin-left:[\s\d\.-]+?', doc['style']).group()
							if margin != margin_norm:
								indent_flag = 1
			if flag_0 == 1:
				groups_1 = sorted(set(groups), key = groups.index)
				if len(groups_1) > 1:
					string_1 = ' '.join(groups_1)
				else:
					string_1 = groups_1[0]
			else:
				string_1 = ''
			if 'colspan' in list(term.attrs.keys()):
				if col_0 == 1:
					col_0 = 2
				if string_1 == '%':
					string_1 = ' %'
				sheet.cell(row = row_0, column = col_0).value = string_1
				if string_1 != '':
					if bold_flag == 0:
						sheet.cell(row = row_0, column = col_0).font = font_1
					else:
						sheet.cell(row = row_0, column = col_0).font = font_2
				else:
					sheet.cell(row = row_0, column = col_0).font = font_1
				col_0 += 1
				sheet.cell(row = row_0, column = col_0).value = ''
				sheet.cell(row = row_0, column = col_0).font = font_1			
			else:
				sheet.cell(row = row_0, column = col_0).value = string_1
				if string_1 != '':
					if bold_flag == 0:
						sheet.cell(row = row_0, column = col_0).font = font_1
					else:
						sheet.cell(row = row_0, column = col_0).font = font_2
				else:
					sheet.cell(row = row_0, column = col_0).font = font_1
			if (col_0 == 1 and indent_flag == 1):
				sheet.cell(row = row_0, column = col_0).alignment = align_1

	nrows = sheet.max_row
	ncols = sheet.max_column
				
	for j in range(1, nrows + 1):	# The first step to clear NONE.
		for k in range(1, ncols + 1):
			if sheet.cell(row = j, column = k).value is None:
				sheet.cell(row = j, column = k).value = ''
	
	for j in range(1, nrows + 1):	# The second step to remove outliers.VERY PERFECT.
		for k in range(2, ncols + 1):
			nut_0 = sheet.cell(row = j, column = k)
			nut_2 = sheet.cell(row = j, column = k + 1)
			if nut_0.value == '$':
				nut_0.value = '$' + nut_2.value
				nut_2.value = ''
			else:
				if ')' in nut_0.value:
					left_count = len(re.findall('\(', nut_0.value))
					right_count = len(re.findall('\)', nut_0.value))
					if left_count < right_count:
						for num in range(1, k):
							nut = sheet.cell(row = j, column = k - num)
							if '(' in nut.value:
								nut.value += nut_0.value
								nut_0.value = ''
								break
				else:
					if (nut_0.value == '%' or nut_0.value == 'ppt'):
						for num in range(1, k):
							nut = sheet.cell(row = j, column = k - num)
							if nut.value != '':
								nut.value += nut_0.value
								nut_0.value = ''
								break

	ncols = typeset(ncols, nrows, sheet)	#	To complete the bracket.
	for j in range(1, nrows + 1):
		for k in range(1, ncols + 1):
			smile = sheet.cell(row = j, column = k)
			if (not re.search('\(\S{1}\)', smile.value) is None and re.sub('\(\S{1}\)', '', smile.value).strip() == ''):
				for u in range(k - 1, 1, -1):
					laugh = sheet.cell(row = j, column = u)
					if laugh.value != '':
						laugh.value += smile.value
						smile.value = ''
						break
	title = 1	#	To get the row in which the heads lie.
	max_count = 0
	for j in range(1, nrows + 1):
		column_count = 0
		for k in range(2, ncols + 1):
			if sheet.cell(row = j, column = k).value != '':
				column_count += 1
		if column_count > max_count:
			max_count = column_count
			title = j

	first_list = []		#	To get to know which rows aren't blank at the first column.
	for j in range(title, nrows + 1):
		if sheet.cell(row = j, column = 1).value != '':
			first_list.append(j)
	
	for k in range(2, ncols + 1):
		for j in range(1, title):
			if sheet.cell(row = j, column = k).value != '':
				dim_flag = 0
				for z in range(j + 1, nrows + 1):
					if sheet.cell(row = z, column = k).value != '':
						dim_flag = 1
						break
				if dim_flag == 0:
					sheet.cell(row = j, column = k + 1).value = sheet.cell(row = j, column = k).value
					sheet.cell(row = j, column = k).value = ''
					break

	in_col = []	#	To get to know which columns aren't blank.
	for j in range(2, ncols + 1):
		flag_1 = 0
		for k in range(title + 1, nrows + 1):
			var = sheet.cell(row = k, column = j).value
			if var != '':
				flag_1 = 1
				break
		if flag_1 == 1:
			in_col.append(j)
	in_col.sort()

	title_cols = []
	for k in range(2, ncols + 1):
		if sheet.cell(row = title, column = k).value != '':
			title_cols.append(k)
	if title_cols == []:
		title_cols = [2]
	for k in title_cols:
		shine_flag = 0
		for j in range(title + 1, nrows + 1):
			if sheet.cell(row = j, column = k).value != '':
				shine_flag = 1
		if shine_flag == 0:
			for num in in_col:
				if num > k:
					break
			for j in range(title + 1, nrows + 1):
				sheet.cell(row = j, column = k).value = sheet.cell(row = j, column = num).value
				sheet.cell(row = j, column = num).value = ''
	
	in_col = []	#	To get to know which columns aren't blank.
	for j in range(2, ncols + 1):
		flag_1 = 0
		for k in range(title + 1, nrows + 1):
			var = sheet.cell(row = k, column = j).value
			if (var != '' and sheet.cell(row = title, column = j).value == ''):
				flag_1 = 1
				break
		if flag_1 == 1:
			in_col.append(j)
	k = 0
	while k < 2:
		ncols = typeset(ncols, nrows, sheet)	#	Adjust the position of columns.
		arrange(ncols, nrows, title, sheet, first_list)
		k += 1
	ncols = typeset(ncols, nrows, sheet)	#	Adjust the position of columns.	
	map = {1 : 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y', 26: 'Z'}	#	Set up the map from the number to the character.
	for k in range(1, ncols + 1):
		max_width = 8.38
		for j in range(1, nrows + 1):
			var = sheet.cell(row = j, column = k).value
			width = 8.38 * (len(var)/8)
			if max_width < width:
				max_width = width
		if k in list(map.keys()):
			sheet.column_dimensions[map[k]].width = max_width
	
	row_0 = title + 1	#	Fill the cells with colors.
	color_cols = []
	while (row_0 < nrows + 1):
		for k in range(1, ncols + 1):
			sheet.cell(row = row_0, column = k).fill = fill_1
		row_0 += 2

w.remove_sheet(w.get_sheet_by_name(w.get_sheet_names()[0]))
if types in ['10-Q', '10-q']:
	file_name = cik + '_' + year + '-' + quart + '-' + types
else:
	file_name = cik + '_' + year + '-' + types
w.save(file_name + '.xlsx')
